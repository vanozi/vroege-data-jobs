"""Export helpers for the kippen registration app."""

from csv import writer
from io import BytesIO, StringIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table
from reportlab.platypus import TableStyle


WEEK_HEADERS = [
    "Dag",
    "Datum",
    "Koppel",
    "Leeftijd",
    "1e soort",
    "2e soort",
    "Dagtotaal",
    "Dode hennen",
    "Buitennest",
    "Eigewicht (g)",
    "Water (ml)",
    "Voer (g)",
    "Opmerkingen",
]


def weekly_calendar_xlsx(
    *,
    flock_week: int,
    week_start: object,
    rows: list[dict[str, object]],
    totals: dict[str, float],
) -> BytesIO:
    """Build an Excel workbook for one weekly laying calendar."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = f"Leeftijdsweek {flock_week}"
    worksheet.append([f"Legkalender leeftijdsweek {flock_week} ({week_start})"])
    worksheet.append([])
    worksheet.append(WEEK_HEADERS)

    for row in rows:
        worksheet.append(_week_export_row(row))

    worksheet.append(
        [
            "Week totaal",
            "",
            "",
            "",
            totals["first_quality_eggs"],
            totals["second_quality_eggs"],
            totals["total_eggs"],
            totals["dead_hens_count"],
            totals["outside_nest_egg_count"],
            _optional_decimal(totals["average_egg_weight_grams"]),
            totals["water_ml"],
            totals["feed_grams"],
            "",
        ]
    )

    worksheet["A1"].font = Font(bold=True, size=14)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[3]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for cell in worksheet[worksheet.max_row]:
        cell.font = Font(bold=True)

    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(
            max_length + 2, 40
        )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def weekly_calendar_pdf(
    *,
    flock_week: int,
    week_start: object,
    rows: list[dict[str, object]],
    totals: dict[str, float],
) -> BytesIO:
    """Build a PDF for one weekly laying calendar."""
    output = BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24,
    )
    styles = getSampleStyleSheet()
    data = [WEEK_HEADERS]
    data.extend(_week_export_row(row) for row in rows)
    data.append(
        [
            "Week totaal",
            "",
            "",
            "",
            totals["first_quality_eggs"],
            totals["second_quality_eggs"],
            totals["total_eggs"],
            totals["dead_hens_count"],
            totals["outside_nest_egg_count"],
            _optional_decimal(totals["average_egg_weight_grams"]),
            totals["water_ml"],
            totals["feed_grams"],
            "",
        ]
    )
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightblue),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("ALIGN", (4, 1), (11, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
            ]
        )
    )
    document.build(
        [
            Paragraph(
                f"Legkalender leeftijdsweek {flock_week} ({week_start})",
                styles["Title"],
            ),
            Spacer(1, 12),
            table,
        ]
    )
    output.seek(0)
    return output


def records_csv(headers: list[str], rows: list[list[object]]) -> BytesIO:
    """Build a UTF-8 CSV file for raw records."""
    output = StringIO()
    csv_writer = writer(output)
    csv_writer.writerow(headers)
    csv_writer.writerows(rows)
    return BytesIO(output.getvalue().encode("utf-8-sig"))


def _week_export_row(row: dict[str, object]) -> list[object]:
    egg_registration = row["egg_registration"]
    feed_water_registration = row["feed_water_registration"]

    return [
        row["weekday"],
        row["date"].isoformat(),
        _flock_name(row),
        _flock_age_label(row),
        _egg_value(egg_registration, "first_quality_eggs"),
        _egg_value(egg_registration, "second_quality_eggs"),
        _egg_value(egg_registration, "total_eggs"),
        row["dead_hens_count"],
        row["outside_nest_egg_count"],
        _optional_decimal(row["average_egg_weight_grams"]),
        _feed_water_value(feed_water_registration, "water_ml"),
        _feed_water_value(feed_water_registration, "feed_grams"),
        _week_notes(egg_registration, feed_water_registration),
    ]


def _egg_value(registration, field_name: str) -> object:
    if registration is None:
        return ""

    return getattr(registration, field_name)


def _feed_water_value(registration, field_name: str) -> object:
    if registration is None:
        return ""

    return getattr(registration, field_name)


def _optional_decimal(value) -> object:
    if value is None:
        return ""

    return value


def _week_notes(egg_registration, feed_water_registration) -> str:
    notes = []
    if egg_registration is not None and egg_registration.notes:
        notes.append(egg_registration.notes)
    if feed_water_registration is not None and feed_water_registration.notes:
        notes.append(feed_water_registration.notes)

    return " | ".join(notes)


def _flock_name(row: dict[str, object]) -> str:
    flock = row.get("flock")
    if flock is None:
        return ""

    return flock.flock_name


def _flock_age_label(row: dict[str, object]) -> str:
    age_context = row.get("flock_age")
    if age_context is None:
        return ""

    return str(age_context["label"])
